#!/usr/bin/python
# Input: party id, cn name, en name, start ip, end ip (same five columns as NSTL layout,
# but column A is Party.partyId instead of serial id). See bulk-ip-upload skill.

import django
import sys
import os
import pandas as pd
import xlsxwriter

os.sys.path.append('../')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'paywall2.settings')
django.setup()

from party.models import IpRange
from common.party_ip_row_check import (
    ip_check_error_row_party_id,
    normalize_excel_party_id,
    validate_party_id_row_for_load,
)


def _read_bulk_party_ip_excel(path):
    """xlrd 2.x cannot read .xlsx; prefer openpyxl for .xlsx."""
    ext = os.path.splitext(path)[1].lower()
    if ext != ".xlsx":
        return pd.read_excel(
            path, index_col=None, usecols="A,B,C,D,E", na_filter=False
        )
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "Reading .xlsx requires openpyxl. Install: pip install 'openpyxl<3'"
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if len(rows) < 2:
        return pd.DataFrame(columns=["party id", "cn name", "en name", "start ip", "end ip"])
    df = pd.DataFrame(rows[1:], columns=[str(c).strip() for c in rows[0]])
    want = ["party id", "cn name", "en name", "start ip", "end ip"]
    for col in want:
        if col not in df.columns:
            raise ValueError("Missing column %r in %s; got %s" % (col, path, list(df.columns)))
    return df[want]


# Begin main program:

IpRangeFilename = sys.argv[1]

data = _read_bulk_party_ip_excel(IpRangeFilename)
output = []
errList = []
toExpire = set()
toIgnore = set()
count = 0
print(len(data))
for index, row in data.iterrows():
    party_id = normalize_excel_party_id(row['party id'])
    cn_name = row['cn name']
    en_name = row['en name']
    start = row['start ip']
    end = row['end ip']
    if count % 10 == 0:
        print(str(count) + '/' + str(len(data)))
    count += 1

    try:
        ok, err_rows, ign, exp, start_n, end_n = validate_party_id_row_for_load(
            party_id, cn_name, en_name, start, end, IpRange)
        errList.extend(err_rows)
        toIgnore.update(ign)
        toExpire.update(exp)
        if ok:
            output.append([party_id, cn_name, en_name, start_n, end_n])
    except Exception as e:
        errList.append(ip_check_error_row_party_id(
            party_id, cn_name, en_name, start, end, getattr(e, 'message', str(e))))
        continue

print('error IP ranges to ignore:' + ','.join(toIgnore))
print('total ' + str(len(toIgnore)))
print('IP ranges need to expire:' + ','.join(toExpire))
print('total ' + str(len(toExpire)))

df = pd.DataFrame(output,
    columns=['party id', 'cn name', 'en name', 'start ip', 'end ip'])

errdf = pd.DataFrame(errList,
    columns=['party id', 'cn name', 'en name', 'start ip', 'end ip', 'error', 'ip range id'])

df.to_excel('bulk_party_ip_to_load.xlsx', engine='xlsxwriter', index=False)
errdf.to_excel('bulk_party_ip_check_error.xlsx', engine='xlsxwriter', index=False)
