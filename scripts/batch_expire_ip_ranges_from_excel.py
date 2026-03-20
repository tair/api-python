#!/usr/bin/python
# Batch-expire IpRange rows listed in an Excel file (by ipRangeId).
#
# Expected: a column headed "ip range id" (as in NSTL_all_ip_check_error.xlsx), or
# "ipRangeId" / "ip_range_id" (case-insensitive, spaces ignored).
#
# Usage (from project root with PYTHONPATH, or cwd=scripts/ like other NSTL scripts):
#   python scripts/batch_expire_ip_ranges_from_excel.py <path_to.xlsx> [--dry-run]
#
# Examples:
#   docker compose run --rm -w /var/www/api-python -e PYTHONPATH=/var/www/api-python \
#     paywall-api-python2 python scripts/batch_expire_ip_ranges_from_excel.py /path/to/expire.xlsx --dry-run

from __future__ import print_function

import argparse
import os
import re
import sys

import django
import pandas as pd

_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _root not in sys.path:
    sys.path.insert(0, _root)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "paywall2.settings")
django.setup()

from django.db import transaction
from django.utils import timezone

from party.models import IpRange


def _read_excel(path):
    """xlrd 2.x does not read .xlsx; load those with openpyxl into a DataFrame."""
    ext = os.path.splitext(path)[1].lower()
    if ext != ".xlsx":
        return pd.read_excel(path, index_col=None, na_filter=False)
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "Reading .xlsx requires openpyxl. Install: pip install 'openpyxl<3'",
            file=sys.stderr,
        )
        sys.exit(1)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows[1:], columns=rows[0])


def _norm_header(name):
    if name is None:
        return ""
    return re.sub(r"\s+", "", str(name).strip().lower())


def _find_id_column(df):
    # Headers like "ip range id", "ipRangeId", "ip_range_id"
    wanted = {"iprangeid", "ip_range_id"}
    for col in df.columns:
        if _norm_header(col) in wanted:
            return col
    return None


def _parse_id(val):
    if val is None:
        return None
    if isinstance(val, basestring) and not str(val).strip():
        return None
    if val == "":
        return None
    try:
        i = int(float(val))
        return i
    except (ValueError, TypeError):
        return None


def main():
    parser = argparse.ArgumentParser(
        description="Set expiredAt=now for IpRange ids listed in an Excel column."
    )
    parser.add_argument("excel_path", help="Path to .xls or .xlsx")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print plan only; do not update the database",
    )
    parser.add_argument(
        "--column",
        default=None,
        help="Force column name for ipRangeId (default: auto-detect)",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.excel_path):
        print("File not found: {}".format(args.excel_path), file=sys.stderr)
        return 1

    df = _read_excel(args.excel_path)
    col = args.column
    if col is None:
        col = _find_id_column(df)
    if col is None or col not in df.columns:
        print(
            "Could not find ip range id column. Columns: {}. "
            "Pass --column 'Your Header'".format(list(df.columns)),
            file=sys.stderr,
        )
        return 1

    ids = []
    seen = set()
    for raw in df[col]:
        rid = _parse_id(raw)
        if rid is None:
            continue
        if rid not in seen:
            seen.add(rid)
            ids.append(rid)

    if not ids:
        print("No ip range ids found in column {!r}".format(col))
        return 0

    existing = list(IpRange.objects.filter(ipRangeId__in=ids))
    by_id = {r.ipRangeId: r for r in existing}
    missing = [i for i in ids if i not in by_id]
    already = [i for i in ids if i in by_id and by_id[i].expiredAt is not None]
    to_expire = [by_id[i] for i in ids if i in by_id and by_id[i].expiredAt is None]

    print("Unique ids in file: {}".format(len(ids)))
    print("Found in DB: {}".format(len(existing)))
    print("Not found (skipped): {}".format(len(missing)))
    if missing:
        print("Missing ids: {}".format(missing[:50] + (["..."] if len(missing) > 50 else [])))
    print("Already expired (skipped): {}".format(len(already)))
    print("To expire now: {}".format(len(to_expire)))

    if args.dry_run:
        print("Dry-run: no database changes.")
        return 0

    if not to_expire:
        print("Nothing to update.")
        return 0

    now = timezone.now()
    pks = [r.ipRangeId for r in to_expire]
    with transaction.atomic():
        updated = IpRange.objects.filter(ipRangeId__in=pks, expiredAt__isnull=True).update(
            expiredAt=now
        )
    print("Expired {} IP range(s).".format(updated))
    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
