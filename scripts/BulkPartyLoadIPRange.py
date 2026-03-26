#!/usr/bin/python
# -*- coding: utf-8 -*-
# Input: party id, cn name, en name, start ip, end ip (from BulkPartyLoadIPCheck output).
# Validates each row with the same rules as BulkPartyLoadIPCheck before inserting.
# If any row fails, writes bulk_party_ip_load_validation_error.xlsx and exits without loading.

import django
import pandas as pd
import sys
import os
import xlsxwriter

os.sys.path.append('../')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'paywall2.settings')
django.setup()

from party.models import Party, IpRange
from common.party_ip_row_check import (
    ip_check_error_row_party_id,
    normalize_excel_party_id,
    validate_party_id_row_for_load,
)


def _read_bulk_party_ip_excel(path):
    ext = os.path.splitext(path)[1].lower()
    if ext != ".xlsx":
        return pd.read_excel(
            path, index_col=None, usecols="A,B,C,D,E", na_filter=False
        )
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "Reading .xlsx requires openpyxl. Install: pip install 'openpyxl<3'"
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if len(rows) < 2:
        return pd.DataFrame(columns=["party id", "cn name", "en name", "start ip", "end ip"])
    df = pd.DataFrame(rows[1:], columns=[str(c).strip() for c in rows[0]])
    want = ["party id", "cn name", "en name", "start ip", "end ip"]
    for col in want:
        if col not in df.columns:
            raise ValueError("Missing column %r in %s; got %s" % (col, path, list(df.columns)))
    return df[want]


IpRangeFilename = sys.argv[1]

data = _read_bulk_party_ip_excel(IpRangeFilename)
output = []
errList = []

count = 0
print(len(data))
for index, row in data.iterrows():
    party_id = normalize_excel_party_id(row['party id'])
    cn_name = row['cn name']
    en_name = row['en name']
    start = row['start ip']
    end = row['end ip']
    if count % 10 == 0:
        print(str(count) + '/' + str(len(data)))
    count += 1
    try:
        ok, err_rows, _, _, start_n, end_n = validate_party_id_row_for_load(
            party_id, cn_name, en_name, start, end, IpRange)
        errList.extend(err_rows)
        if ok:
            output.append([party_id, cn_name, en_name, start_n, end_n])
    except Exception as e:
        errList.append(ip_check_error_row_party_id(
            party_id, cn_name, en_name, start, end, getattr(e, 'message', str(e))))
        continue

if errList:
    errdf = pd.DataFrame(
        errList,
        columns=['party id', 'cn name', 'en name', 'start ip', 'end ip', 'error', 'ip range id'])
    errdf.to_excel('bulk_party_ip_load_validation_error.xlsx', engine='xlsxwriter', index=False)
    print('Validation failed; wrote bulk_party_ip_load_validation_error.xlsx — no IP ranges were loaded.')
    sys.exit(1)

loadingErr = []
for party_id, cn_name, en_name, start, end in output:
    print(party_id)
    try:
        p = Party.objects.get(partyId=party_id)
        IpRange.objects.create(partyId=p, start=start, end=end)
    except Exception as e:
        loadingErr.append([party_id, cn_name, en_name, start, end, e.message])
        continue

if loadingErr:
    print('error saved in bulk_party_ip_load_runtime_error.xlsx')
    diffdf = pd.DataFrame(
        loadingErr,
        columns=['party id', 'cn name', 'en name', 'start ip', 'end ip', 'error'])
    diffdf.to_excel('bulk_party_ip_load_runtime_error.xlsx', engine='xlsxwriter', index=False)
