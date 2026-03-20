#!/usr/bin/python
# input file format: serial id, cn name, en name, start ip, end ip
#
# Validates each row with the same rules as NSTLLoadIPCheck before inserting.
# If any row fails, writes NSTL_all_ip_load_validation_error.xlsx and exits without loading.

import django
import pandas as pd
import sys
import os
import xlsxwriter

# format: serial id,cn name,en name,start ip,end ip
os.sys.path.append('../')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'paywall2.settings')
django.setup()

from party.models import Party, IpRange, PartyAffiliation, Country
from common.nstl_ip_row_check import (
    ip_check_error_row,
    normalize_excel_serial_id,
    validate_nstl_row_for_load,
)


def _read_nstl_ip_excel(path):
    """xlrd 2.x cannot read .xlsx; NSTL_check outputs are xlsx (xlsxwriter)."""
    ext = os.path.splitext(path)[1].lower()
    if ext != ".xlsx":
        return pd.read_excel(
            path, index_col=None, usecols="A,B,C,D,E", na_filter=False
        )
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "Reading .xlsx requires openpyxl. Install: pip install 'openpyxl<3'"
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if len(rows) < 2:
        return pd.DataFrame(columns=["serial id", "cn name", "en name", "start ip", "end ip"])
    df = pd.DataFrame(rows[1:], columns=[str(c).strip() for c in rows[0]])
    want = ["serial id", "cn name", "en name", "start ip", "end ip"]
    for col in want:
        if col not in df.columns:
            raise ValueError("Missing column %r in %s; got %s" % (col, path, list(df.columns)))
    return df[want]


# Begin main program:

IpRangeFilename = sys.argv[1]

data = _read_nstl_ip_excel(IpRangeFilename)
output = []
errList = []

count = 0
print(len(data))
for index, row in data.iterrows():
    serialId = normalize_excel_serial_id(row['serial id'])
    cn_name = row['cn name']
    en_name = row['en name']
    start = row['start ip']
    end = row['end ip']
    if count % 10 == 0:
        print(str(count) + '/' + str(len(data)))
    count += 1
    try:
        ok, err_rows, _, _, start_n, end_n = validate_nstl_row_for_load(
            serialId, cn_name, en_name, start, end, IpRange)
        errList.extend(err_rows)
        if ok:
            output.append([serialId, cn_name, en_name, start_n, end_n])
    except Exception as e:
        errList.append(ip_check_error_row(
            serialId, cn_name, en_name, start, end, getattr(e, 'message', str(e))))
        continue

if errList:
    errdf = pd.DataFrame(
        errList,
        columns=['serial id', 'cn name', 'en name', 'start ip', 'end ip', 'error', 'ip range id'])
    errdf.to_excel('NSTL_all_ip_load_validation_error.xlsx', engine='xlsxwriter', index=False)
    print('Validation failed; wrote NSTL_all_ip_load_validation_error.xlsx — no IP ranges were loaded.')
    sys.exit(1)

loadingErr = []
for serialId, cn_name, en_name, start, end in output:
    print(serialId)
    try:
        p = None
        if not Party.objects.filter(serialId=serialId).exists():
            country = Country.objects.get(countryId=127)
            p = Party.objects.create(
                partyType='organization', serialId=serialId, display=True,
                name=en_name, country=country)
            NSTL = Party.objects.get(partyId=31772)
            PartyAffiliation.objects.create(childPartyId=p, parentPartyId=NSTL)
        else:
            p = Party.objects.get(serialId=serialId)
        IpRange.objects.create(partyId=p, start=start, end=end)
    except Exception as e:
        loadingErr.append([serialId, cn_name, en_name, start, end, e.message])
        continue

if loadingErr:
    print('error saved in NSTL_all_ip_load_runtime_error.xlsx')
    diffdf = pd.DataFrame(
        loadingErr,
        columns=['serial id', 'cn name', 'en name', 'start ip', 'end ip', 'error'])
    diffdf.to_excel('NSTL_all_ip_load_runtime_error.xlsx', engine='xlsxwriter', index=False)
