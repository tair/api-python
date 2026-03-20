#!/usr/bin/env python3
"""
Filter NSTL_all_ip_check_error.xlsx for review: drop exact DB matches, same-institution
\"contained\" pairs, and non-overlap errors; keep real conflicts and cross-institution
containment.

**Dependencies:** CPython 3 + openpyxl only — **no Django, no DB, no app code.** Safe to run on
the host or in a one-off ``python:3.x-slim`` container with the repo mounted (see
``bulk_ip_load_pipeline.py`` and cursor-skills bulk-nstl-ip-load docs).

Requires: pip install openpyxl

Usage (from project root):
  python3 scripts/filter_nstl_ip_check_error_overlap_only.py \\
    [--input NSTL_all_ip_check_error.xlsx] \\
    [--output NSTL_all_ip_check_error_overlap_only.xlsx]

- \"ip range exists\" — dropped (identical to DB).
- \"ip range contained\" — NSTLLoadIPCheck emits two consecutive rows (spreadsheet row +
  matching DB row). If both \"serial id\" values match, the new range lies inside an existing
  range for the same institution (redundant); both rows are dropped, like exact matches.
  If serial ids differ, the pair is kept (new range inside another institution's range).
- Validation errors (too large, private, invalid) stay only in the full file.
"""
from __future__ import print_function

import argparse
import os
import sys

EXACT_MATCH = "ip range exists"
CONTAINED = "ip range contained"
OVERLAP_TYPES = frozenset(
    (
        CONTAINED,
        "ip range contain existing in same institution",
        "ip range contain existing in different institution",
        "ip range overlap in same institution",
        "ip range overlap in different institution",
    )
)


def main():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        print(
            "Missing dependency: openpyxl. Install with: python3 -m pip install openpyxl",
            file=sys.stderr,
        )
        return 1

    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    default_in = os.path.join(root, "NSTL_all_ip_check_error.xlsx")
    default_out = os.path.join(root, "NSTL_all_ip_check_error_overlap_only.xlsx")

    p = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    p.add_argument("--input", default=default_in, help="Full error workbook from NSTLLoadIPCheck")
    p.add_argument("--output", default=default_out, help="Filtered workbook for review")
    args = p.parse_args()

    if not os.path.isfile(args.input):
        print("Input not found: {}".format(args.input), file=sys.stderr)
        return 1

    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print("Input workbook is empty.", file=sys.stderr)
        return 1

    header = list(rows[0])
    try:
        err_idx = header.index("error")
    except ValueError:
        print("No 'error' column in header: {}".format(header), file=sys.stderr)
        return 1

    hl = len(header)

    def norm_err(val):
        if val is None:
            return ""
        return str(val).strip()

    def norm_serial(val):
        if val is None:
            return ""
        s = str(val).strip()
        if s.lower() == "nan":
            return ""
        return s

    def pad_row(row):
        padded = list(row[:hl])
        while len(padded) < hl:
            padded.append(None)
        return padded

    def msg_at(row):
        if len(row) <= err_idx:
            return ""
        return norm_err(row[err_idx])

    data_rows = list(rows[1:])
    n = len(data_rows)
    idx = 0
    total = n
    removed_exact = 0
    removed_same_inst_contained = 0
    skipped_other = 0
    kept_overlap = 0
    body = []

    while idx < n:
        row = data_rows[idx]
        msg = msg_at(row)

        if msg == EXACT_MATCH:
            removed_exact += 1
            idx += 1
            continue

        if (
            msg == CONTAINED
            and idx + 1 < n
            and msg_at(data_rows[idx + 1]) == CONTAINED
        ):
            row2 = data_rows[idx + 1]
            if norm_serial(row[0]) == norm_serial(row2[0]):
                removed_same_inst_contained += 2
                idx += 2
                continue
            body.append(pad_row(row))
            body.append(pad_row(row2))
            kept_overlap += 2
            idx += 2
            continue

        if msg in OVERLAP_TYPES:
            body.append(pad_row(row))
            kept_overlap += 1
            idx += 1
            continue

        skipped_other += 1
        idx += 1

    out_wb = Workbook(write_only=True)
    out_ws = out_wb.create_sheet()
    out_ws.append(header)
    for r in body:
        out_ws.append(r)
    out_wb.save(args.output)

    contained_pairs = removed_same_inst_contained // 2
    print(
        "Rows read (excl. header): {} | "
        "removed exact match ('{}'): {} | "
        "removed same-inst '{}' rows: {} ({} consecutive row pairs) | "
        "skipped other errors (see full file): {} | "
        "overlap review rows written: {}".format(
            total,
            EXACT_MATCH,
            removed_exact,
            CONTAINED,
            removed_same_inst_contained,
            contained_pairs,
            skipped_other,
            kept_overlap,
        )
    )
    print("Wrote: {}".format(args.output))
    return 0


if __name__ == "__main__":
    sys.exit(main())
