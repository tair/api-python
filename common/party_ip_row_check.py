# Party-keyed Excel row validation (size, private, overlap, exact match) for bulk IP load
# by Party.partyId. Same behavioral rules as common.nstl_ip_row_check.validate_nstl_row_for_load,
# but "same institution" uses partyId instead of serialId.
# Used by scripts/BulkPartyLoadIPCheck.py and scripts/BulkPartyLoadIPRange.py (Python 2).
from __future__ import absolute_import

import numbers

try:
    basestring
except NameError:
    basestring = str

from netaddr import IPAddress

from common.common import (
    exact_match_exists,
    get_overlapping_ranges,
    ip2long,
    isIpRangePrivate,
    validateIpRangeSize,
)


def normalize_excel_party_id(val):
    """
    Integer Party.partyId after pandas/openpyxl infers float (e.g. 354790.0).
    Returns int or None if missing/invalid.
    """
    if val is None or val == '':
        return None
    try:
        if val != val:
            return None
    except TypeError:
        pass
    if isinstance(val, float):
        i = int(val)
        if val == i:
            return i
        return None
    if isinstance(val, numbers.Integral) and not isinstance(val, bool):
        return int(val)
    s = str(val).strip()
    if not s or s.lower() == 'nan':
        return None
    try:
        return int(s)
    except (TypeError, ValueError):
        pass
    try:
        f = float(s)
        if f == int(f):
            return int(f)
    except (TypeError, ValueError, OverflowError):
        pass
    return None


def ip_check_error_row_party_id(party_id, cn_name, en_name, start, end, error, ip_range_id=''):
    """Seven columns for errdf; first column is party id (string for Excel)."""
    pid = '' if party_id is None else str(party_id).strip()
    return [pid, cn_name, en_name, start, end, error, ip_range_id]


def _coerce_ip_cell(val):
    """Normalize Excel cell to a stripped string (handles float/int from pandas/openpyxl)."""
    if val is None:
        return ''
    try:
        if isinstance(val, float) and val != val:
            return ''
    except Exception:
        pass
    if isinstance(val, basestring):
        return val.strip()
    return str(val).strip()


def validate_party_id_row_for_load(party_id_int, cn_name, en_name, start, end, IpRange):
    """
    Same rules as validate_nstl_row_for_load, but "same institution" uses Party.partyId
    instead of serialId. party_id_int must be a valid Party primary key.

    Returns (ok_to_load, err_rows, to_ignore_add, to_expire_add, start_norm, end_norm).
    """
    from party.models import Party

    to_ignore_add = set()
    to_expire_add = set()
    err_rows = []

    pid_display = str(party_id_int) if party_id_int is not None else ''

    start = _coerce_ip_cell(start)
    end = _coerce_ip_cell(end)
    if party_id_int is None:
        err_rows.append(ip_check_error_row_party_id(
            '', cn_name, en_name, start, end, 'party id missing or invalid'))
        return False, err_rows, to_ignore_add, to_expire_add, start, end

    try:
        Party.objects.get(partyId=party_id_int)
    except Party.DoesNotExist:
        err_rows.append(ip_check_error_row_party_id(
            pid_display, cn_name, en_name, start, end, 'party id not found'))
        return False, err_rows, to_ignore_add, to_expire_add, start, end

    if start and not end:
        end = start
    if end and not start:
        start = end
    if not start or not end:
        err_rows.append(ip_check_error_row_party_id(
            pid_display, cn_name, en_name, start, end, 'ip range invalid'))
        return False, err_rows, to_ignore_add, to_expire_add, start, end

    if not validateIpRangeSize(start, end):
        err_rows.append(ip_check_error_row_party_id(
            pid_display, cn_name, en_name, start, end, 'ip range too large'))
        return False, err_rows, to_ignore_add, to_expire_add, start, end

    if isIpRangePrivate(start, end):
        err_rows.append(ip_check_error_row_party_id(
            pid_display, cn_name, en_name, start, end, 'ip range private'))
        return False, err_rows, to_ignore_add, to_expire_add, start, end

    if start.startswith('255') or end.startswith('255'):
        err_rows.append(ip_check_error_row_party_id(
            pid_display, cn_name, en_name, start, end, 'ip range invalid'))
        return False, err_rows, to_ignore_add, to_expire_add, start, end

    try:
        IPAddress(start)
        IPAddress(end)
    except Exception:
        err_rows.append(ip_check_error_row_party_id(
            pid_display, cn_name, en_name, start, end, 'ip range invalid'))
        return False, err_rows, to_ignore_add, to_expire_add, start, end

    if exact_match_exists(start, end, IpRange):
        existing = get_overlapping_ranges(start, end, IpRange)
        try:
            s_long = ip2long(start)
            e_long = ip2long(end)
        except Exception:
            s_long, e_long = None, None
        for ipRange in existing:
            if s_long is not None and ipRange.startLong == s_long and ipRange.endLong == e_long:
                err_rows.append(ip_check_error_row_party_id(
                    pid_display, cn_name, en_name, start, end, 'ip range exists'))
                err_rows.append(ip_check_error_row_party_id(
                    ipRange.partyId_id, '', ipRange.partyId.name,
                    ipRange.start, ipRange.end, 'ip range exists', ipRange.ipRangeId))
                to_ignore_add.add(str(ipRange.ipRangeId))
        if not err_rows:
            err_rows.append(ip_check_error_row_party_id(
                pid_display, cn_name, en_name, start, end, 'ip range exists'))
        return False, err_rows, to_ignore_add, to_expire_add, start, end

    overlapping = get_overlapping_ranges(start, end, IpRange)
    overlap = False
    for ipRange in overlapping:
        range_party_id = ipRange.partyId_id
        range_names = ipRange.partyId.name
        range_ids_str = str(ipRange.ipRangeId)
        start_val = IPAddress(start)
        end_val = IPAddress(end)
        range_start = IPAddress(ipRange.start)
        range_end = IPAddress(ipRange.end)
        same_inst = party_id_int == range_party_id

        if start_val >= range_start and end_val <= range_end:
            overlap = True
            err_rows.append(ip_check_error_row_party_id(
                pid_display, cn_name, en_name, start, end, 'ip range contained'))
            err_rows.append(ip_check_error_row_party_id(
                range_party_id, '', range_names, ipRange.start, ipRange.end, 'ip range contained', range_ids_str))
            to_ignore_add.add(str(ipRange.ipRangeId))
        elif start_val <= range_start and end_val >= range_end:
            overlap = True
            if same_inst:
                err_rows.append(ip_check_error_row_party_id(
                    pid_display, cn_name, en_name, start, end, 'ip range contain existing in same institution'))
                err_rows.append(ip_check_error_row_party_id(
                    range_party_id, '', range_names, ipRange.start, ipRange.end,
                    'ip range contain existing in same institution', range_ids_str))
            else:
                err_rows.append(ip_check_error_row_party_id(
                    pid_display, cn_name, en_name, start, end, 'ip range contain existing in different institution'))
                err_rows.append(ip_check_error_row_party_id(
                    range_party_id, '', range_names, ipRange.start, ipRange.end,
                    'ip range contain existing in different institution', range_ids_str))
            to_expire_add.add(str(ipRange.ipRangeId))
        else:
            overlap = True
            if same_inst:
                err_rows.append(ip_check_error_row_party_id(
                    pid_display, cn_name, en_name, start, end, 'ip range overlap in same institution'))
                err_rows.append(ip_check_error_row_party_id(
                    range_party_id, '', range_names, ipRange.start, ipRange.end,
                    'ip range overlap in same institution', range_ids_str))
            else:
                err_rows.append(ip_check_error_row_party_id(
                    pid_display, cn_name, en_name, start, end, 'ip range overlap in different institution'))
                err_rows.append(ip_check_error_row_party_id(
                    range_party_id, '', range_names, ipRange.start, ipRange.end,
                    'ip range overlap in different institution', range_ids_str))
            to_expire_add.add(str(ipRange.ipRangeId))

    if overlap:
        return False, err_rows, to_ignore_add, to_expire_add, start, end

    return True, [], to_ignore_add, to_expire_add, start, end
 
 
 def _read_bulk_party_ip_excel(path):
     """xlrd 2.x cannot read .xlsx; prefer openpyxl for .xlsx."""
     import os
     import pandas as pd
 
     ext = os.path.splitext(path)[1].lower()
     if ext != ".xlsx":
         return pd.read_excel(
             path, index_col=None, usecols="A,B,C,D,E", na_filter=False
         )
     try:
         from openpyxl import load_workbook
     except ImportError:
         raise ImportError(
             "Reading .xlsx requires openpyxl. Install: pip install 'openpyxl<3'"
         )
     wb = load_workbook(path, read_only=True, data_only=True)
     ws = wb.active
     rows = [list(r) for r in ws.iter_rows(values_only=True)]
     wb.close()
     if len(rows) < 2:
         return pd.DataFrame(columns=["party id", "cn name", "en name", "start ip", "end ip"])
     df = pd.DataFrame(rows[1:], columns=[str(c).strip() for c in rows[0]])
     want = ["party id", "cn name", "en name", "start ip", "end ip"]
     for col in want:
         if col not in df.columns:
             raise ValueError("Missing column %r in %s; got %s" % (col, path, list(df.columns)))
     return df[want]
